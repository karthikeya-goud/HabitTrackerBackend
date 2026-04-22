def calculate_completion(value, target, tasktype):
    if tasktype == "checkbox":
        return 100 if value == 1 else 0
    return min((value / target) * 100, 100)


def calculate_score(completion, max_score):
    return (completion / 100) * max_score


from django.utils import timezone
from datetime import timedelta


def get_period_range(frequency):
    now = timezone.now()

    if frequency == "daily":
        start = now.replace(hour=0, minute=0, second=0)
    elif frequency == "weekly":
        start = now - timedelta(days=now.weekday())
        start = start.replace(hour=0, minute=0, second=0)
    elif frequency == "monthly":
        start = now.replace(day=1, hour=0, minute=0, second=0)

    return start, now


def is_today(last_task_date):
    now = timezone.localdate()
    return now == timezone.localdate(last_task_date)


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.utils import timezone
from datetime import timedelta


def generate_excel(tasklogs):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Habits"

    center = Alignment(horizontal="center", vertical="center")

    # -----------------------------
    # 1. Get last 7 days (you can change this)
    # -----------------------------
    today = timezone.now().date()
    dates = [today - timedelta(days=i) for i in range(6, -1, -1)]

    date_headers = [d.strftime("%b %d") for d in dates]

    # -----------------------------
    # 2. Collect unique tasks
    # -----------------------------
    tasks = {}
    for log in tasklogs:
        tasks[log.task.id] = log.task

    tasks = list(tasks.values())

    # -----------------------------
    # 3. Header row
    # -----------------------------
    headers = ["Habit"] + date_headers
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = center

    # -----------------------------
    # 4. Fill rows
    # -----------------------------
    for task in tasks:
        row = [task.title]

        for date in dates:
            log = next(
                (
                    l
                    for l in tasklogs
                    if l.task.id == task.id and l.created_at.date() == date
                ),
                None,
            )

            if not log:
                row.append("-")
                continue

            # checkbox
            if task.type == "checkbox":
                row.append("✔" if log.value == 1 else "✘")

            # slider / target
            else:
                if task.target_value:
                    row.append(
                        f"{log.value}/{task.target_value} {task.unit if task.unit not in [None,'steps'] else ''}"
                    )
                else:
                    row.append(str(log.value))

        ws.append(row)

        # center align row
        for col in range(1, len(row) + 1):
            ws.cell(row=ws.max_row, column=col).alignment = center

    # -----------------------------
    # 5. Column widths
    # -----------------------------
    ws.column_dimensions["A"].width = 25
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 12

    ws.freeze_panes = "B2"

    return wb
